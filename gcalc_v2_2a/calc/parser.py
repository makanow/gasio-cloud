from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, Any
import openpyxl
import pandas as pd

@dataclass
class EntryData:
    basic: Dict[str, Any]
    sales: pd.DataFrame
    land: pd.DataFrame
    assets: pd.DataFrame
    road: pd.DataFrame


def load_entry_xlsx(xlsx_bytes: bytes) -> EntryData:
    wb = openpyxl.load_workbook(filename=bytes(xlsx_bytes), data_only=True)

    ws = wb['01_基本情報']
    basic: Dict[str, Any] = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        val = ws.cell(r, 3).value
        if key:
            basic[str(key)] = val

    def read_sheet(name: str) -> pd.DataFrame:
        ws = wb[name]
        rows = list(ws.values)
        header = rows[0]
        data = rows[1:]
        df = pd.DataFrame(data, columns=header)
        df = df.dropna(how='all')
        return df

    sales = read_sheet('02_販売量')
    land = read_sheet('03_土地')
    assets = read_sheet('04_償却資産')
    road = read_sheet('05_道路占用料')

    return EntryData(basic=basic, sales=sales, land=land, assets=assets, road=road)
