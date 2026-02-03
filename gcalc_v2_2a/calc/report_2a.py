from __future__ import annotations
import io
from typing import Dict
import openpyxl


def build_report_2a(template_bytes: bytes, computed: Dict[str, float]) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    mapping_amount = {
        'raw_cost': 'E9',
        'labor_cost': 'E10',
        'repair_cost': 'E11',
        'property_tax': 'E12',
        'business_tax': 'E13',
        'road_occupancy': 'E14',
        'depreciation': 'E15',
        'other_expenses': 'E16',
        'subtotal': 'E17',
        'remuneration': 'E18',
        'corp_tax': 'E19',
        'local_corp_tax': 'E20',
        'inhabitant_tax': 'E21',
        'total_cost': 'E22',
        'sales_m3': 'E24',
        'unit_yen_per_m3': 'E25',
    }
    for k, addr in mapping_amount.items():
        ws[addr].value = computed.get(k, 0)

    total = float(computed.get('total_cost', 0) or 0)
    comp_keys = [
        ('raw_cost','G9'),('labor_cost','G10'),('repair_cost','G11'),('property_tax','G12'),
        ('business_tax','G13'),('road_occupancy','G14'),('depreciation','G15'),('other_expenses','G16'),
        ('subtotal','G17'),('remuneration','G18'),('corp_tax','G19'),('local_corp_tax','G20'),('inhabitant_tax','G21'),
        ('total_cost','G22'),
    ]
    for k, addr in comp_keys:
        v = float(computed.get(k, 0) or 0)
        ws[addr].value = (v / total * 100) if total else 0

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
